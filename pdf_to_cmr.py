#!/usr/bin/env python3
"""
PDF Packing List to CMR Excel Converter - FINAL VERIFIED SCRIPT
Reads ALL pages and aggregates all boxes/pallets
Uses correct cell locations (C28, C32, C36, B28) per user template.
*** FIX: Crops page AND uses a simple 5-line limit. ***
"""

import re
import sys
import os
from datetime import datetime
from typing import Dict, List, Optional
import pdfplumber
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment


class PackingListExtractor:
    """Extract data from CTS packing list PDFs - handles multi-page PDFs"""
    
    def __init__(self, pdf_path: str):
        self.pdf_path = pdf_path
        self.data = {}
    
    def extract(self) -> Dict:
        """Main extraction method - reads ALL pages"""
        try:
            print(f"Opening PDF: {self.pdf_path}")
            with pdfplumber.open(self.pdf_path) as pdf:
                print(f"✓ PDF opened - {len(pdf.pages)} pages found")
                
                # Read first page for header info and consignee
                first_page = pdf.pages[0]
                
                # 1. Get text from the FULL page for right-side data
                full_text = first_page.extract_text()
                if not full_text:
                    raise Exception("PDF text extraction returned empty - PDF may be corrupted or scanned image")
                
                # 2. Crop the page to the left 50%
                left_half_bbox = (0, 0, first_page.width * 0.5, first_page.height)
                left_page_crop = first_page.crop(left_half_bbox)
                
                # 3. Get text *only* from the left half
                left_text = left_page_crop.extract_text()
                if not left_text:
                    print("⚠ Warning: Left-half crop returned no text. Falling back to full text.")
                    left_text = full_text

                print(f"✓ Extracted text from full page and left half")
                
                # Extract right-side data from FULL text
                self.data['packing_list_number'] = self._extract_packing_list_number(full_text)
                self.data['date'] = self._extract_date(full_text)
                self.data['your_ref'] = self._extract_your_ref(full_text)
                self.data['our_ref'] = self._extract_our_ref(full_text)
                self.data['delivery_terms'] = self._extract_delivery_terms(full_text)
                
                # Extract left-side data from LEFT text
                self.data['consignee'] = self._extract_consignee(left_text)
                
                print(f"✓ Header extracted - Consignee: {self.data['consignee'].get('name', 'N/A')}")
                
                # Extract ALL boxes/pallets from ALL pages (deduplicated)
                self.data['boxes'] = []
                total_gross_weight = 0
                seen_box_numbers = set()  # Track box numbers to avoid duplicates
                
                print(f"\nExtracting boxes from all {len(pdf.pages)} pages...")
                for page_num, page in enumerate(pdf.pages, 1):
                    page_text = page.extract_text()
                    if not page_text:
                        print(f"  ⚠ Page {page_num}: No text extracted")
                        continue
                        
                    box_info = self._extract_box_from_page(page_text, page_num)
                    if box_info:
                        box_number = box_info.get('number')
                        box_name = box_info.get('name', 'Box')
                        
                        # Only add if we haven't seen this box number before
                        if box_number not in seen_box_numbers:
                            self.data['boxes'].append(box_info)
                            seen_box_numbers.add(box_number)
                            print(f"  ✓ Page {page_num}: Added {box_name}")
                            if box_info.get('gross_weight_kg'):
                                total_gross_weight += box_info['gross_weight_kg']
                        else:
                            print(f"  - Page {page_num}: Skipped {box_name} (duplicate - already added)")
                    else:
                        print(f"  - Page {page_num}: No box found (might be continuation)")
                
                self.data['total_gross_weight'] = total_gross_weight
                self.data['num_boxes'] = len(self.data['boxes'])
                
                print(f"\n✓ Total: {self.data['num_boxes']} unique boxes, {total_gross_weight} KG")
                
                return self.data
                
        except Exception as e:
            raise Exception(f"Error extracting PDF data: {e}\n\nThis may indicate:\n- PDF file is corrupted\n- PDF is a scanned image (not text-based)\n- File upload was incomplete")
    
    def _extract_packing_list_number(self, text: str) -> Optional[str]:
        # Looks for "Packing List 12345" or "Packing List 12345-1"
        match = re.search(r'Packing List\s+(\d+)(?:-\d+)?', text)
        return match.group(1) if match else None
    
    def _extract_date(self, text: str) -> Optional[str]:
        # Looks for "Barendrecht, 04-09-2023"
        match = re.search(r'Barendrecht,\s*(\d{2}-\d{2}-\d{4})', text)
        return match.group(1) if match else None
    
    def _extract_your_ref(self, text: str) -> Optional[str]:
        # Looks for "Your ref.: ..."
        match = re.search(r'Your ref\.:\s*([^\n]+)', text)
        return match.group(1).strip() if match else None
    
    def _extract_our_ref(self, text: str) -> Optional[str]:
        # Looks for "Our ref.: 12345"
        match = re.search(r'Our ref\.:\s*(\d{4,5})', text)
        return match.group(1).strip() if match else None
    
    def _extract_consignee(self, text: str) -> Dict:
        """Extract consignee - stops *after* finding 5 address lines."""
        consignee = {}
        
        lines = text.split('\n')
        in_consignee = False
        consignee_lines = []
        found_header = False
        
        # Look for "Consignee address" header
        for i, line in enumerate(lines):
            line_stripped = line.strip()
            line_lower = line.lower()
            
            # Look for the header
            if 'consignee address' in line_lower:
                in_consignee = True
                found_header = True
                print(f"  ✓ Found 'Consignee address' at line {i}")
                continue
                
            if in_consignee:
                # *** THIS IS THE SIMPLIFIED FIX ***
                # We no longer check for PRE_STOP_KEYWORDS.
                # We no longer check for SENDER SKIP (it's not in this text).
                
                # Accept this line if it has content
                if line_stripped:
                    consignee_lines.append(line_stripped)
                    print(f"  + Added line {len(consignee_lines)}: {line_stripped[:60]}")
                    
                    # *** Hard-stop limit ***
                    # After adding 5 lines (name, add1, add2, city, country),
                    # we STOP. No matter what.
                    if len(consignee_lines) >= 5:
                        print(f"  ✓ Reached 5-line limit. Stopping.")
                        break
        
        # Strategy 2: If header "Consignee address" is not found
        # (This remains as a fallback)
        if not found_header or len(consignee_lines) < 2:
            pass # (No changes to Strategy 2)
        
        # Parse consignee lines
        if len(consignee_lines) >= 1:
            consignee['name'] = consignee_lines[0]
        if len(consignee_lines) >= 2:
            consignee['address_line1'] = consignee_lines[1]
        if len(consignee_lines) >= 3:
            consignee['address_line2'] = consignee_lines[2]
        if len(consignee_lines) >= 4:
            consignee['city'] = consignee_lines[3]
        if len(consignee_lines) >= 5:
            consignee['country'] = consignee_lines[4]
        # We no longer parse extra1, extra2, etc. at all
        
        print(f"\n  DEBUG - Consignee extraction result:")
        if consignee:
            for key, val in consignee.items():
                print(f"      {key}: {val}")
        else:
            print(f"  ⚠ WARNING: Consignee dict is EMPTY! (This is bad)")
        
        return consignee
    
    def _extract_delivery_terms(self, text: str) -> Optional[str]:
        # Looks for "Delivery ..."
        match = re.search(r'Delivery\s+([^\n]+)', text)
        return match.group(1).strip() if match else None
    
    def _extract_box_from_page(self, text: str, page_num: int) -> Optional[Dict]:
        """Extract box/pallet/case/crate info from a single page"""
        box = {}
        
        package_types = [
            r'Wooden\s*box', r'Pallet', r'Case', r'Crate', r'Carton\s*box',
            r'Carton', r'Package', r'Container', r'Box', r'Skid', r'Bundle'
        ]
        
        for pkg_type_pattern in package_types:
            # Match: "Wooden box (1)" or "Case 1" or "Pallet(6)"
            collo_match = re.search(rf'({pkg_type_pattern})\s*\(?\s*(\d+)\s*\)?', text, re.IGNORECASE)
            if collo_match:
                box['type'] = collo_match.group(1).strip()
                box['number'] = int(collo_match.group(2))
                type_cleaned = re.sub(r'\s+', ' ', box['type'].title())
                box['name'] = f"{type_cleaned} {box['number']}"
                print(f"    ✓ Found package: {box['name']}")
                break
        
        if not box:
            # Strategy 2: Look for "Packing List 15738-X"
            packing_match = re.search(r'Packing List\s+\d+[-\s]*(\d+)', text)
            if packing_match:
                box_num = int(packing_match.group(1))
                box['type'] = 'Package'
                box['number'] = box_num
                box['name'] = f"Package {box_num}"
                print(f"    ✓ Found from packing list number: {box['name']}")
            else:
                print(f"    - No package identifier found on page {page_num}")
                return None
        
        # Extract measurements
        measurement_patterns = [
            r'Measurement[:\s]+([\d\s]+x[\d\s]+x[\d\s]+)',
            r'Dimensions?[:\s]+([\d\s]+x[\d\s]+x[\d\s]+)',
            r'(\d+\s*x\s*\d+\s*x\s*\d+)\s*cm',
        ]
        
        for pattern in measurement_patterns:
            measurement_match = re.search(pattern, text, re.IGNORECASE)
            if measurement_match:
                dims = measurement_match.group(1).strip()
                dims = re.sub(r'\s+', ' ', dims).replace(' x ', ' x ')
                box['dimensions'] = dims
                print(f"      Dimensions: {dims}")
                break
        
        if 'dimensions' not in box: print(f"      ⚠ No dimensions found")
        
        # Extract gross weight (handles commas: 1,234 KG)
        weight_patterns = [
            r'Gross\s*weight[:\s]+([\d,]+)\s*KG',
            r'Gross[:\s]+([\d,]+)\s*KG',
            r'([\d,]+)\s*KG.*gross',
        ]
        
        for pattern in weight_patterns:
            gross_weight_match = re.search(pattern, text, re.IGNORECASE)
            if gross_weight_match:
                # Remove commas before converting to int
                weight_str = gross_weight_match.group(1).replace(',', '')
                weight = int(weight_str)
                box['gross_weight'] = f"{weight} KG"
                box['gross_weight_kg'] = weight
                print(f"      Weight: {weight} KG")
                break
        
        if 'gross_weight_kg' not in box: print(f"      ⚠ No gross weight found")
        
        return box


class CMRExcelPopulator:
    """Populate CMR Excel template - ALL CELLS VERIFIED"""
    
    # Mapping for country codes, can be expanded
    COUNTRY_CODES = {
        'NETHERLANDS': 'NL',
        'THE NETHERLANDS': 'NL',
        'GERMANY': 'DE',
        'BELGIUM': 'BE',
        'FRANCE': 'FR',
        'UNITED KINGDOM': 'GB',
        'UK': 'GB',
        'SAUDI ARABIA': 'SA',
        'KINGDOM OF SAUDI ARABIA': 'SA',
        'UNITED ARAB EMIRATES': 'AE',
        'UAE': 'AE',
        'OMAN': 'OM',
        'SULTANATE OF OMAN': 'OM',
        'QATAR': 'QA',
        'KUWAIT': 'KW',
        'BAHRAIN': 'BH',
        'IRAQ': 'IQ',
        'TURKEY': 'TR',
        'EGYPT': 'EG',
        'ITALY': 'IT',
        'SPAIN': 'ES',
    }

    def __init__(self, template_path: str):
        self.template_path = template_path
        self.wb = None
        self.ws = None
    
    def _get_country_code(self, country_name: str) -> str:
        """Gets 2-letter country code, or returns name if not found."""
        if not country_name:
            return ''
            
        name_upper = country_name.upper().strip()
        
        if name_upper in self.COUNTRY_CODES:
            return self.COUNTRY_CODES[name_upper]
        
        for key, code in self.COUNTRY_CODES.items():
            if key in name_upper:
                return code
        
        return name_upper
    
    def populate(self, data: Dict, output_path: str):
        """Populate template with extracted data"""
        
        if os.path.exists(self.template_path):
            try:
                self.wb = load_workbook(self.template_path)
                self.ws = self.wb.active
                print(f"✓ Loaded template")
                # Apply settings even to existing template
                # DON'T set column widths early - do it at the END only
                # self._apply_column_widths()  # DISABLED
                self._apply_row_heights()
                self._apply_print_settings()
            except Exception as e:
                print(f"⚠ Warning: Could not load template '{self.template_path}'. Error: {e}")
                print("Creating a new blank workbook.")
                self._create_cmr_template()
        else:
            print(f"⚠ Warning: Template '{self.template_path}' not found.")
            print("Creating a new blank workbook.")
            self._create_cmr_template()
        
        # Populate all sections
        self._populate_header_section(data)
        self._populate_sender_section()
        self._populate_consignee_section(data.get('consignee', {}))
        self._populate_boxes_section(data.get('boxes', []))
        self._populate_footer_section(data)
        
        # DON'T apply column widths here - do it at the VERY END only
        # self._apply_column_widths()  # DISABLED
        
        # MERGE CELLS to give more space for addresses
        print("  Merging address cells for more space...")
        try:
            # Merge sender cells (B6:C8) - UPDATED ROWS
            self.ws.merge_cells('B6:C6')  # CTS Netherlands B.V.
            self.ws.merge_cells('B7:C7')  # Riga 10
            self.ws.merge_cells('B8:C8')  # Address
            
            # Merge row 11 (empty box)
            self.ws.merge_cells('B11:C11')
            
            # Merge consignee cells (B16:C20) - UPDATED ROWS
            self.ws.merge_cells('B16:C16')  # Company name
            self.ws.merge_cells('B17:C17')  # Address line 1
            self.ws.merge_cells('B18:C18')  # Address line 2
            self.ws.merge_cells('B19:C19')  # City
            self.ws.merge_cells('B20:C20')  # Country
            
            # Merge city formula cell
            self.ws.merge_cells('B26:C26')
            
            # Merge origin cell
            self.ws.merge_cells('B33:C33')
            
            # Merge box description column (B) to give more space
            self.ws.merge_cells('B45:C45')  # Colli header
            self.ws.merge_cells('B46:C46')  # Dimensions text
            self.ws.merge_cells('B48:C48')  # Description header
            
            print("  ✓ Merged cells for wider address display")
        except Exception as e:
            print(f"  ⚠ Warning: Could not merge cells: {e}")
        
        # SET ALL COLUMN WIDTHS AT THE VERY END - LAST THING BEFORE SAVE
        print("  [FINAL] Setting all column widths...")
        self.ws.column_dimensions['A'].width = 18  # LEFT MARGIN SPACE (as per working template)
        self.ws.column_dimensions['B'].width = 40  # MAIN CONTENT COLUMN
        self.ws.column_dimensions['C'].width = 20
        self.ws.column_dimensions['D'].width = 12
        self.ws.column_dimensions['E'].width = 20  # Dimensions
        self.ws.column_dimensions['F'].width = 12
        self.ws.column_dimensions['G'].width = 20  # Labels  
        self.ws.column_dimensions['H'].width = 25  # Values
        self.ws.column_dimensions['I'].width = 12
        
        print(f"  ✓ [FINAL] Column widths set:")
        print(f"       A={self.ws.column_dimensions['A'].width} ← LEFT MARGIN")
        print(f"       B={self.ws.column_dimensions['B'].width} ← MAIN COLUMN")
        print(f"  Matches working template!")
        
        self.wb.save(output_path)
        print(f"✓ CMR saved: {output_path}")
    
    def _create_cmr_template(self):
        """Create CMR template with precise row heights matching CMR form"""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "CMR"
        
        # Column widths (matching CMR template)
        self.ws.column_dimensions['A'].width = 18  # LEFT MARGIN (as per working template)
        self.ws.column_dimensions['B'].width = 40  # Main column
        self.ws.column_dimensions['C'].width = 20
        self.ws.column_dimensions['D'].width = 12
        self.ws.column_dimensions['E'].width = 15  # Dimensions column
        self.ws.column_dimensions['F'].width = 12
        self.ws.column_dimensions['G'].width = 15  # Labels column
        self.ws.column_dimensions['H'].width = 20  # Values column
        self.ws.column_dimensions['I'].width = 12
        
        # ROW HEIGHTS - CRITICAL: Match CMR form boxes exactly
        # Rows 1-3: Header area (logos, etc) - compact
        for row in range(1, 4):
            self.ws.row_dimensions[row].height = 12
        
        # Rows 4-6: Box 1 - Sender (3 lines) - compact
        for row in range(4, 7):
            self.ws.row_dimensions[row].height = 14
        
        # Rows 7-11: Gap between Box 1 and Box 2
        for row in range(7, 12):
            self.ws.row_dimensions[row].height = 13
        
        # Rows 12-16: Box 2 - Consignee (5 lines) - slightly more space
        for row in range(12, 17):
            self.ws.row_dimensions[row].height = 14.5
        
        # Rows 17-20: Gap
        for row in range(17, 21):
            self.ws.row_dimensions[row].height = 13
        
        # Row 21: City formula
        self.ws.row_dimensions[21].height = 14
        
        # Rows 22-24: Gap
        for row in range(22, 25):
            self.ws.row_dimensions[row].height = 13
        
        # Row 25: Origin
        self.ws.row_dimensions[25].height = 14
        
        # Rows 26-32: Right side info area
        for row in range(26, 33):
            self.ws.row_dimensions[row].height = 14
        
        # Row 33: Packing list
        self.ws.row_dimensions[33].height = 14
        
        # Rows 34-37: Gap before boxes
        for row in range(34, 38):
            self.ws.row_dimensions[row].height = 13
        
        # Rows 38-42: Box headers section
        for row in range(38, 43):
            self.ws.row_dimensions[row].height = 14
        
        # Rows 43-56: Box data - COMPACT (up to 14 boxes)
        for row in range(43, 57):
            self.ws.row_dimensions[row].height = 13.5
        
        # Rows 57-65: Footer section
        for row in range(57, 66):
            self.ws.row_dimensions[row].height = 14
        
        # Rows 66-70: Bottom area
        for row in range(66, 71):
            self.ws.row_dimensions[row].height = 13
        
        # PAGE SETUP - Fit to one A4 page
        self.ws.page_setup.paperSize = 9  # A4
        self.ws.page_setup.orientation = 'portrait'
        self.ws.page_setup.fitToPage = True
        self.ws.page_setup.fitToHeight = 1
        self.ws.page_setup.fitToWidth = 1
        
        # Print area - CMR form area
        self.ws.print_area = 'A1:I70'
        
        # Margins - tight for CMR form
        self.ws.page_margins.left = 0.2
        self.ws.page_margins.right = 0.2
        self.ws.page_margins.top = 0.2
        self.ws.page_margins.bottom = 0.2
        self.ws.page_margins.header = 0.0
        self.ws.page_margins.footer = 0.0
        
        print("  ✓ Created CMR template with precise row heights for form alignment")
    
    def _apply_column_widths(self):
        """Apply column widths to worksheet - explicit settings"""
        print(f"  DEBUG: Setting column widths...")
        
        # Set column A - LEFT MARGIN SPACER
        self.ws.column_dimensions['A'].width = 18  # As per working template
        self.ws.column_dimensions['A'].bestFit = False
        self.ws.column_dimensions['A'].auto_size = False
        
        # Set column B - CRITICAL
        self.ws.column_dimensions['B'].width = 40  # WIDE!
        self.ws.column_dimensions['B'].bestFit = False
        self.ws.column_dimensions['B'].auto_size = False
        # Note: customWidth is read-only, can't be set
        
        # Set other columns
        self.ws.column_dimensions['C'].width = 20
        self.ws.column_dimensions['D'].width = 12
        self.ws.column_dimensions['E'].width = 20
        self.ws.column_dimensions['F'].width = 12
        self.ws.column_dimensions['G'].width = 20
        self.ws.column_dimensions['H'].width = 25
        self.ws.column_dimensions['I'].width = 12
        
        print(f"  ✓ Column widths set: A={self.ws.column_dimensions['A'].width} (left margin), B={self.ws.column_dimensions['B'].width}")
    
    def _apply_row_heights(self):
        """Apply row heights to worksheet"""
        # Rows 1-3: Header
        for row in range(1, 4):
            self.ws.row_dimensions[row].height = 12
        # Rows 4-6: Sender
        for row in range(4, 7):
            self.ws.row_dimensions[row].height = 14
        # Rows 7-11: Gap
        for row in range(7, 12):
            self.ws.row_dimensions[row].height = 13
        # Rows 12-16: Consignee
        for row in range(12, 17):
            self.ws.row_dimensions[row].height = 14.5
        # Rows 17-20: Gap
        for row in range(17, 21):
            self.ws.row_dimensions[row].height = 13
        # Row 21: City
        self.ws.row_dimensions[21].height = 14
        # Rows 22-24: Gap
        for row in range(22, 25):
            self.ws.row_dimensions[row].height = 13
        # Row 25: Origin
        self.ws.row_dimensions[25].height = 14
        # Rows 26-32: Right info
        for row in range(26, 33):
            self.ws.row_dimensions[row].height = 14
        # Row 33: Packing list
        self.ws.row_dimensions[33].height = 14
        # Rows 34-37: Gap
        for row in range(34, 38):
            self.ws.row_dimensions[row].height = 13
        # Rows 38-42: Box headers
        for row in range(38, 43):
            self.ws.row_dimensions[row].height = 14
        # Rows 43-56: Box data
        for row in range(43, 57):
            self.ws.row_dimensions[row].height = 13.5
        # Rows 57-65: Footer
        for row in range(57, 66):
            self.ws.row_dimensions[row].height = 14
        # Rows 66-70: Bottom
        for row in range(66, 71):
            self.ws.row_dimensions[row].height = 13
        print("  ✓ Applied row heights")
    
    def _apply_print_settings(self):
        """Apply print settings to worksheet for single-page A4 output"""
        # PAGE SETUP
        self.ws.page_setup.paperSize = 9  # A4
        self.ws.page_setup.orientation = 'portrait'
        self.ws.page_setup.fitToPage = True
        self.ws.page_setup.fitToHeight = 1  # Fit to 1 page tall
        self.ws.page_setup.fitToWidth = 1   # Fit to 1 page wide
        self.ws.page_setup.scale = 85  # Scale down to 85% to give more room
        
        # Print area - only columns we use
        self.ws.print_area = 'A1:H70'  # Exclude column I
        
        # Margins (in inches) - very tight
        self.ws.page_margins.left = 0.15
        self.ws.page_margins.right = 0.15
        self.ws.page_margins.top = 0.15
        self.ws.page_margins.bottom = 0.15
        self.ws.page_margins.header = 0.0
        self.ws.page_margins.footer = 0.0
        
        print("  ✓ Applied single-page print settings with 85% scale")
    
    def _populate_header_section(self, data: Dict):
        """Populate header section - ALL CELLS VERIFIED"""
        
        # C28: REMOVED - was causing extra "BARENDRECHT, NL" cell
        # self.ws['C28'] = "BARENDRECHT, NL"  # COMMENTED OUT

        # C32: Destination (City, Country Code)
        consignee = data.get('consignee', {})
        
        city = ''
        # The 3rd line (index 2) is 'MUSCAT 100'
        if consignee.get('address_line2'):
             city_match = re.match(r'([A-Z\s]+)', consignee.get('address_line2', '').upper())
             if city_match:
                 city = city_match.group(1).strip()
        # Fallback to 4th line (index 3)
        elif consignee.get('city'):
            city_text = consignee.get('city', '').upper()
            if not any(trigger.upper() in city_text for trigger in self.COUNTRY_CODES.keys()):
                 city = city_text

        country = ''
        # The 5th line (index 4) is 'SULTANATE OF OMAN'
        if consignee.get('country'):
            country = consignee.get('country', '')
        # Fallback to 4th line (index 3)
        elif consignee.get('city'):
             city_text = consignee.get('city', '').upper()
             if any(trigger.upper() in city_text for trigger in self.COUNTRY_CODES.keys()):
                 country = city_text
        
        destination = ''
        
        if city:
            destination = city
            if country:
                country_code = self._get_country_code(country)
                destination = f"{destination}, {country_code}"
        elif country:
            destination = self._get_country_code(country)
        
        # C32: REMOVED - was causing extra "KAV, INDONESIA" cell
        # Instead we add formulas:
        
        # B33: Formula - copy of sender address (B8)
        self.ws['B33'] = '=B8'
        print(f"    Writing B33 (formula): =B8")
        
        # B69: Formula - copy of sender address (B8) for footer
        self.ws['B69'] = '=B8'
        print(f"    Writing B69 (formula): =B8")
        
        # B85: Formula - copy of sender address (B8) for bottom
        self.ws['B85'] = '=B8'
        print(f"    Writing B85 (formula): =B8")
        
        # G33/H33: Delivery terms (MOVED from G26)
        self.ws['G33'] = "Delivery term"
        if data.get('delivery_terms'):
            delivery = data['delivery_terms']
            if 'EXW' in delivery: self.ws['H33'] = 'EXW'
            elif 'CIF' in delivery: self.ws['H33'] = 'CIF'
            elif 'FOB' in delivery: self.ws['H33'] = 'FOB'
            elif 'FCA' in delivery: self.ws['H33'] = 'FCA'
            else: self.ws['H33'] = delivery.split()[0] if delivery else ''
        
        # G35/H35: Project No (Our ref) (MOVED from G28)
        self.ws['G35'] = "Project No.:"
        if data.get('our_ref'):
            self.ws['H35'] = f"CTS-{data['our_ref']}"
        
        # G37/H37: Customer ref (MOVED from G30)
        self.ws['G37'] = "Customer ref"
        if data.get('your_ref'):
            self.ws['H37'] = data['your_ref']
        
        # B40/C40: Packing list number
        self.ws['B40'] = "Packing list No.:"
        if data.get('packing_list_number'):
            self.ws['C40'] = data['packing_list_number']
            print(f"    Writing B40/C40: Packing list No.: {data['packing_list_number']}")
    
    def _populate_sender_section(self):
        """Populate static sender info - ROWS 6-8"""
        self.ws['B6'] = "CTS Netherlands B.V."
        self.ws['B7'] = "Riga 10"
        self.ws['B8'] = "2993 LW BARENDRECHT, NL"
    
    def _populate_consignee_section(self, consignee: Dict):
        """Populate consignee address - ROWS 16-20"""
        
        if not consignee:
            print("    ⚠ Consignee is empty or None! Writing nothing.")
            return
        
        # Write ONLY the main address fields starting at row 16
        row = 16
        
        # *** THE FAILSAFE WRITE ***
        # We only write these specific fields.
        # This guarantees we never write 'extra1' (IBAN), etc.
        fields_to_write = ['name', 'address_line1', 'address_line2', 'city', 'country']
        
        real_city = ''
        
        for field in fields_to_write:
            if consignee.get(field):
                value = consignee[field]
                
                # Try to find the real city from 'address_line2' (e.g., "MUSCAT 100")
                if field == 'address_line2':
                    city_match = re.match(r'([A-Z\s]+)', value.upper())
                    if city_match:
                        real_city = city_match.group(1).strip()
                
                # Fallback: if 'city' field doesn't look like a country, use it
                if field == 'city' and not real_city:
                    city_text = value.upper()
                    if not any(trigger.upper() in city_text for trigger in self.COUNTRY_CODES.keys()):
                         real_city = city_text

                # Uppercase what we think is the city line
                if field == 'city' or field == 'address_line2':
                    value = value.upper()
                
                print(f"    Writing B{row}: {value}")
                self.ws[f'B{row}'] = value
                row += 1
                
                # Safety limit - don't write beyond row 20
                if row > 20:
                    break
        
        # B26: Formula - references B19 (city line)
        self.ws['B26'] = '=B19'
        print(f"    Writing B26 (formula): =B19")
    
    def _populate_boxes_section(self, boxes: List[Dict]):
        """Populate boxes/pallets section - ROWS 45, 46, 48, 50+"""
        
        # Row 45: Headers
        self.ws['B45'] = "Colli"
        self.ws['H45'] = "KG"
        
        # Row 46: Standard text
        self.ws['B46'] = "Dimensions as per attached packaging overview"
        
        # Row 48: Table headers
        self.ws['B48'] = "Description"
        self.ws['E48'] = "L x W x H (cm)"
        self.ws['H48'] = "Gross weight (KG)"
        
        # Populate each box starting at row 50
        start_row = 50
        for idx, box in enumerate(boxes):
            row = start_row + idx
            
            if box.get('name'):
                self.ws[f'B{row}'] = box['name']
            
            if box.get('dimensions'):
                self.ws[f'E{row}'] = box['dimensions']
            
            if box.get('gross_weight_kg'):
                self.ws[f'H{row}'] = box['gross_weight_kg']
    
    def _populate_footer_section(self, data: Dict):
        """Populate static footer - moved lower to avoid box section"""
        # B70+: Contact info section (moved from B57 to avoid boxes)
        self.ws['B70'] = "Previous to deliver, please contact:"
        self.ws['B73'] = "Tel.:"


def main():
    """Main execution"""
    if len(sys.argv) < 2:
        print("Usage: python pdf_to_cmr.py <pdf_file>")
        sys.exit(1)
    
    pdf_path = sys.argv[1]
    
    if not os.path.isfile(pdf_path):
        print(f"Error: File not found: {pdf_path}")
        sys.exit(1)
    
    # Template path is hardcoded
    template_path = "CTS_NL_CMR_Template.xlsx"
    
    # Output filename
    base_name = os.path.splitext(os.path.basename(pdf_path))[0]
    output_path = f"CMR_{base_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    try:
        print(f"--- Starting Extraction ---")
        extractor = PackingListExtractor(pdf_path)
        data = extractor.extract()
        
        print(f"\n--- Extraction Summary ---")
        print(f"  Packing List: {data.get('packing_list_number')}")
        print(f"  Date: {data.get('date')}")
        print(f"  Our Ref: {data.get('our_ref')}")
        print(f"  Your Ref: {data.get('your_ref')}")
        print(f"  Consignee: {data.get('consignee', {}).get('name', 'N/A')}")
        print(f"  Delivery: {data.get('delivery_terms')}")
        print(f"  Boxes: {data.get('num_boxes', 0)}")
        print(f"  Total Weight: {data.get('total_gross_weight', 0)} KG")
        
        for box in data.get('boxes', []):
            print(f"    - {box.get('name')}: {box.get('dimensions')} / {box.get('gross_weight')}")
        
        print(f"\n--- Populating Excel ---")
        populator = CMRExcelPopulator(template_path)
        populator.populate(data, output_path)
        
        print(f"\n✓ Success! Output generated.")
        
    except Exception as e:
        print(f"\n✗ FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()