#!/usr/bin/env python3
"""
Example/Test Script for PDF to CMR Converter
Demonstrates how the tool works with sample data
"""

from openpyxl import Workbook
from datetime import datetime
import os


def create_sample_cmr_template():
    """Create a sample CMR template for testing"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "CMR"
    
    # Header
    ws['A1'] = "CTS Netherlands B.V."
    ws['A2'] = "CMR Document - International Road Transport"
    
    # Sender Information
    ws['A4'] = "SENDER:"
    ws['A5'] = "CTS Netherlands B.V."
    ws['A6'] = "Riga 10"
    ws['A7'] = "2993 LW Barendrecht"
    ws['A8'] = "The Netherlands"
    ws['A9'] = "VAT: NL820895374B01"
    
    # Document Reference Fields (labels)
    ws['F4'] = "Packing List No:"
    ws['F5'] = "Date:"
    ws['F6'] = "Your Reference:"
    ws['F7'] = "Our Reference:"
    
    # Data will be filled in these cells:
    ws['G4'] = "[Packing List]"
    ws['G5'] = "[Date]"
    ws['G6'] = "[Your Ref]"
    ws['G7'] = "[Our Ref]"
    
    # Consignee Section
    ws['A11'] = "CONSIGNEE:"
    ws['A12'] = "[Company Name]"
    ws['A13'] = "[Address Line 1]"
    ws['A14'] = "[Address Line 2]"
    ws['A15'] = "[Phone/Fax]"
    
    # Shipping Information
    ws['A18'] = "SHIPPING INFORMATION:"
    ws['A19'] = "Delivery Terms:"
    ws['C19'] = "[Delivery Terms]"
    ws['A20'] = "Case Info:"
    ws['C20'] = "[Case Info]"
    ws['A21'] = "Measurements:"
    ws['C21'] = "[Measurements]"
    ws['A22'] = "Gross Weight:"
    ws['C22'] = "[Gross Weight]"
    ws['A23'] = "Country of Origin:"
    ws['C23'] = "[Country of Origin]"
    
    # Items Header
    ws['A26'] = "ITEMS / GOODS:"
    ws['A27'] = "Description"
    ws['E27'] = "Article No"
    ws['G27'] = "HS Code"
    ws['I27'] = "Quantity"
    
    # Sample data will be added starting from row 28
    
    # Save template
    output_path = "CTS_NL_CMR_Template.xlsx"
    wb.save(output_path)
    print(f"✓ Created sample CMR template: {output_path}")
    
    return output_path


def create_sample_data():
    """Create sample data structure (simulating PDF extraction)"""
    
    sample_data = {
        'packing_list_number': '15880-1',
        'date': '22-07-2025',
        'your_ref': '01/NIPO/19465',
        'our_ref': '5523',
        'consignee': {
            'name': 'Dohat Al Khaleej LLC',
            'address_line1': 'PO BOX 503, PC 133 AL KHUWAIR',
            'address_line2': 'SULTANATE OF OMAN',
            'tel': '+968 24052867',
            'fax': '+968 24054165'
        },
        'delivery_terms': 'EXW Barendrecht (NL)',
        'case_info': 'Case 1 of 1 / 2.63 m3',
        'measurement': '160 x 160 x 103 cm',
        'gross_weight': '287 KG',
        'country_of_origin': 'NL',
        'items': [
            {
                'description': 'CTS60 or CTS70 IFR wiper seal XPE 31.2m per roll 350mm wide, tnk T-9803',
                'article_number': '60WS350XPE',
                'hs_code': '40169990',
                'quantity': '4 Rolls (31.2m each)'
            },
            {
                'description': 'CTS60 joint splice wiper 350 Esther PU 2065 AS, 780x250x1, tnk T-9803',
                'article_number': '60WSJS350PU',
                'hs_code': '40169990',
                'quantity': '48 PCS'
            },
            {
                'description': 'L-bar rim SS304L, P150, holes 12x20, tnk T-9803',
                'article_number': 'BLR150448',
                'hs_code': '73089098',
                'quantity': '210 PCS'
            },
            {
                'description': 'CTS60 spacer bushing PA-66 Diam. Ø10.3/ Ø15 x L=10, tnk T-9803',
                'article_number': '60SB10PA66',
                'hs_code': '40169990',
                'quantity': '750 PCS'
            }
        ]
    }
    
    return sample_data


def demonstrate_conversion():
    """Demonstrate the conversion process with sample data"""
    
    print("\n" + "="*70)
    print("CTS PDF to CMR Converter - Demonstration")
    print("="*70 + "\n")
    
    # Step 1: Create sample template
    print("Step 1: Creating sample CMR template...")
    template_path = create_sample_cmr_template()
    print()
    
    # Step 2: Show sample data (simulating PDF extraction)
    print("Step 2: Sample data (as if extracted from PDF):")
    sample_data = create_sample_data()
    print(f"  Packing List: {sample_data['packing_list_number']}")
    print(f"  Date: {sample_data['date']}")
    print(f"  Reference: {sample_data['our_ref']}")
    print(f"  Consignee: {sample_data['consignee']['name']}")
    print(f"  Items: {len(sample_data['items'])} items")
    print()
    
    # Step 3: Populate template
    print("Step 3: Populating CMR template...")
    try:
        from pdf_to_cmr import CMRExcelPopulator
        
        populator = CMRExcelPopulator(template_path)
        output_path = f"CMR_EXAMPLE_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        populator.populate(sample_data, output_path)
        
        print(f"✓ CMR document created: {output_path}")
        print()
        
        print("="*70)
        print("SUCCESS! Example CMR document has been created.")
        print("="*70)
        print()
        print("Files created:")
        print(f"  1. {template_path} (Sample template)")
        print(f"  2. {output_path} (Generated CMR document)")
        print()
        print("You can now:")
        print("  - Open the generated CMR Excel file to see the result")
        print("  - Compare it with the original packing list PDF")
        print("  - Use this template for your actual conversions")
        print()
        
    except ImportError:
        print("⚠ Note: pdf_to_cmr.py module not found")
        print("This is just a demonstration of the data structure")
        print()
        print("What would happen with real data:")
        print("  ✓ PDF packing list would be parsed")
        print("  ✓ Data would be extracted automatically")
        print("  ✓ CMR template would be populated")
        print("  ✓ Excel file would be saved")
        print()
    
    except Exception as e:
        print(f"Error during demonstration: {e}")
        import traceback
        traceback.print_exc()
    
    print("\nTo use the actual converter:")
    print("  python pdf_to_cmr.py 5523")
    print("  python pdf_to_cmr_gui.py")
    print()


def show_usage_examples():
    """Show practical usage examples"""
    
    print("\n" + "="*70)
    print("Practical Usage Examples")
    print("="*70 + "\n")
    
    examples = [
        {
            'title': 'Convert single packing list by number',
            'command': 'python pdf_to_cmr.py 5523',
            'description': 'Looks for Packing_List_5523.pdf and converts it'
        },
        {
            'title': 'Convert with full file path',
            'command': 'python pdf_to_cmr.py "C:/Documents/Packing_List_5523.pdf"',
            'description': 'Converts specific PDF file'
        },
        {
            'title': 'Use the GUI (easiest)',
            'command': 'python pdf_to_cmr_gui.py',
            'description': 'Opens graphical interface for easy conversion'
        },
        {
            'title': 'Batch process multiple PDFs',
            'command': 'python batch_convert.py ./packing_lists ./output',
            'description': 'Converts all PDFs in folder'
        },
        {
            'title': 'PowerShell (Windows)',
            'command': '.\\convert_pdf_to_cmr.ps1 -Input 5523',
            'description': 'Windows-specific script'
        }
    ]
    
    for i, example in enumerate(examples, 1):
        print(f"{i}. {example['title']}")
        print(f"   Command: {example['command']}")
        print(f"   Description: {example['description']}")
        print()
    
    print("="*70 + "\n")


if __name__ == "__main__":
    # Run demonstration
    demonstrate_conversion()
    
    # Show usage examples
    show_usage_examples()
    
    print("For more information, see:")
    print("  - README.md (Full documentation)")
    print("  - QUICKSTART.md (Quick start guide)")
    print()
